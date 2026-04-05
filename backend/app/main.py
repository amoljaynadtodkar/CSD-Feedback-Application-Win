from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
from typing import Optional, List
import os
import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook

from .database import init_database, get_db, verify_admin, IMAGES_DIR
from .models import (
    LoginRequest,
    LoginResponse,
    CategoryCreate,
    CategoryResponse,
    ProductResponse,
    DemandCreate,
    DemandResponse,
    DemandUpdate,
    FeedbackCreate,
    FeedbackResponse,
)

app = FastAPI()
security = HTTPBasic()

app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r"(https?://(localhost|127\.0\.0\.1)(:\d+)?|file://.*)",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup_event():
    init_database()
    # Ensure images directory exists before mounting for static access
    os.makedirs(IMAGES_DIR, exist_ok=True)


def verify_auth(credentials: HTTPBasicCredentials = Depends(security)):
    if not verify_admin(credentials.username, credentials.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return credentials.username


@app.post("/login", response_model=LoginResponse)
def login(request: LoginRequest):
    if verify_admin(request.username, request.password):
        return LoginResponse(success=True, message="Login successful")
    return LoginResponse(success=False, message="Invalid credentials")


@app.post(
    "/categories", response_model=CategoryResponse, dependencies=[Depends(verify_auth)]
)
def create_category(category: CategoryCreate):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM categories WHERE name = ?", (category.name,))
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="Category already exists")

        cursor.execute("INSERT INTO categories (name) VALUES (?)", (category.name,))
        conn.commit()
        cursor.execute("SELECT * FROM categories WHERE id = ?", (cursor.lastrowid,))
        return CategoryResponse(**dict(cursor.fetchone()))


@app.get(
    "/categories",
    response_model=List[CategoryResponse],
    dependencies=[Depends(verify_auth)],
)
def get_categories():
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM categories ORDER BY name")
        return [CategoryResponse(**dict(row)) for row in cursor.fetchall()]


@app.delete("/categories/{category_id}", dependencies=[Depends(verify_auth)])
def delete_category(category_id: int):
    with get_db() as conn:
        cursor = conn.cursor()

        # Check if category is being used by any products
        cursor.execute("SELECT name FROM categories WHERE id = ?", (category_id,))
        category_row = cursor.fetchone()
        if not category_row:
            raise HTTPException(status_code=404, detail="Category not found")

        category_name = category_row[0]
        cursor.execute(
            "SELECT COUNT(*) FROM products WHERE category = ?", (category_name,)
        )
        product_count = cursor.fetchone()[0]

        if product_count > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot delete category. It is being used by {product_count} product(s)",
            )

        cursor.execute("DELETE FROM categories WHERE id = ?", (category_id,))
        conn.commit()

    return {"success": True}


@app.post(
    "/products", response_model=ProductResponse, dependencies=[Depends(verify_auth)]
)
async def create_product(
    code: str = Form(...),
    name: str = Form(...),
    category: str = Form(...),
    file: Optional[UploadFile] = File(None),
):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM products WHERE code = ?", (code,))
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="Product code already exists")

        if file is not None and file.filename:
            allowed_extensions = {".jpg", ".jpeg", ".png"}
            file_ext = os.path.splitext(file.filename)[1].lower()

            if file_ext not in allowed_extensions:
                raise HTTPException(
                    status_code=400, detail="Invalid file type. Only JPG and PNG allowed."
                )

            if file.size > 5 * 1024 * 1024:
                raise HTTPException(status_code=400, detail="File too large. Maximum 5MB.")

            image_filename = f"{code}{file_ext}"
            image_path = os.path.join(IMAGES_DIR, image_filename)

            with open(image_path, "wb") as buffer:
                content = await file.read()
                buffer.write(content)
        else:
            image_filename = "placeholder.svg"

        cursor.execute(
            "INSERT INTO products (code, name, category, image_path) VALUES (?, ?, ?, ?)",
            (code, name, category, image_filename),
        )
        conn.commit()
        cursor.execute("SELECT * FROM products WHERE id = ?", (cursor.lastrowid,))
        return ProductResponse(**dict(cursor.fetchone()))


@app.post("/products/bulk", dependencies=[Depends(verify_auth)])
async def bulk_add_products(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed.")

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the Excel file. Ensure it is a valid .xlsx file.")

    added = []
    skipped = []
    errors = []

    with get_db() as conn:
        cursor = conn.cursor()

        for sheet_name in wb.sheetnames:
            category = sheet_name.strip()
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find column indices from header row
            header = [str(cell).strip().upper() if cell else "" for cell in rows[0]]
            code_idx = None
            name_idx = None
            for i, h in enumerate(header):
                if "INDEX" in h and "NO" in h:
                    code_idx = i
                elif "NOMENCLATURE" in h:
                    name_idx = i

            if code_idx is None or name_idx is None:
                errors.append(f"Sheet '{category}': Missing 'INDEX NO' or 'Nomenclature' column headers.")
                continue

            # Ensure category exists
            cursor.execute("SELECT id FROM categories WHERE name = ?", (category,))
            if not cursor.fetchone():
                cursor.execute("INSERT INTO categories (name) VALUES (?)", (category,))

            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    code = str(row[code_idx]).strip() if row[code_idx] else ""
                    name = str(row[name_idx]).strip() if row[name_idx] else ""

                    if not code or not name:
                        continue

                    cursor.execute("SELECT id FROM products WHERE code = ?", (code,))
                    if cursor.fetchone():
                        skipped.append(f"{code} - {name} (already exists)")
                        continue

                    cursor.execute(
                        "INSERT INTO products (code, name, category, image_path) VALUES (?, ?, ?, ?)",
                        (code, name, category, "placeholder.svg"),
                    )
                    added.append(f"{code} - {name}")
                except Exception as e:
                    errors.append(f"Sheet '{category}', row {row_num}: {str(e)}")

        conn.commit()

    wb.close()
    return {
        "added": len(added),
        "skipped": len(skipped),
        "errors": len(errors),
        "details": {
            "added": added,
            "skipped": skipped,
            "errors": errors,
        },
    }


@app.get(
    "/products",
    response_model=List[ProductResponse],
    dependencies=[Depends(verify_auth)],
)
def get_products():
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM products")
        return [ProductResponse(**dict(row)) for row in cursor.fetchall()]


@app.get(
    "/products/public",
    response_model=List[ProductResponse],
)
def get_products_public():
    """Public endpoint for customers to browse products"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM products")
        return [ProductResponse(**dict(row)) for row in cursor.fetchall()]


@app.get("/products/categories", dependencies=[Depends(verify_auth)])
def get_product_categories():
    """Get categories for product form dropdown"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM categories ORDER BY name")
        return [row[0] for row in cursor.fetchall()]


@app.get("/products/categories/public")
def get_product_categories_public():
    """Get categories for customer-facing dropdowns (no auth required)"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM categories ORDER BY name")
        return [row[0] for row in cursor.fetchall()]


@app.delete("/products/all", dependencies=[Depends(verify_auth)])
def delete_all_products():
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT image_path FROM products WHERE image_path != 'placeholder.svg' AND image_path IS NOT NULL")
        images = [row[0] for row in cursor.fetchall()]
        cursor.execute("DELETE FROM products")
        conn.commit()

    for img in images:
        try:
            os.remove(os.path.join(IMAGES_DIR, img))
        except FileNotFoundError:
            pass

    return {"success": True, "deleted": len(images)}


@app.delete("/products/{product_id}", dependencies=[Depends(verify_auth)])
def delete_product(product_id: int):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT image_path FROM products WHERE id = ?", (product_id,))
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Product not found")

        image_path = row[0]
        cursor.execute("DELETE FROM products WHERE id = ?", (product_id,))
        conn.commit()

    if image_path:
        try:
            os.remove(os.path.join(IMAGES_DIR, image_path))
        except FileNotFoundError:
            pass

    return {"success": True}


@app.post("/products/bulk-images", dependencies=[Depends(verify_auth)])
async def bulk_add_images(files: List[UploadFile] = File(...)):
    allowed_extensions = {".jpg", ".jpeg", ".png"}
    matched = []
    not_found = []
    errors = []

    with get_db() as conn:
        cursor = conn.cursor()

        for upload_file in files:
            filename = upload_file.filename
            # Strip directory path if browser sends it (webkitdirectory)
            basename = os.path.basename(filename)
            name_without_ext, ext = os.path.splitext(basename)
            ext = ext.lower()

            if ext not in allowed_extensions:
                errors.append(f"{basename}: unsupported file type ({ext})")
                continue

            product_code = name_without_ext.strip()
            cursor.execute("SELECT id, image_path FROM products WHERE code = ?", (product_code,))
            row = cursor.fetchone()

            if not row:
                not_found.append(f"{basename}: no product with code '{product_code}'")
                continue

            product_id = row[0]
            old_image = row[1]

            # Remove old image if it's not placeholder
            if old_image and old_image != "placeholder.svg":
                try:
                    os.remove(os.path.join(IMAGES_DIR, old_image))
                except FileNotFoundError:
                    pass

            image_filename = f"{product_code}{ext}"
            image_path = os.path.join(IMAGES_DIR, image_filename)

            try:
                content = await upload_file.read()
                with open(image_path, "wb") as buffer:
                    buffer.write(content)

                cursor.execute(
                    "UPDATE products SET image_path = ? WHERE id = ?",
                    (image_filename, product_id),
                )
                matched.append(f"{product_code} ({basename})")
            except Exception as e:
                errors.append(f"{basename}: {str(e)}")

        conn.commit()

    return {
        "matched": len(matched),
        "not_found": len(not_found),
        "errors": len(errors),
        "details": {
            "matched": matched,
            "not_found": not_found,
            "errors": errors,
        },
    }


@app.post("/demands", response_model=DemandResponse)
def create_demand(demand: DemandCreate):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO demands (category, product_name, product_code, new_description, quantity, required_by, name, contact_number) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                demand.category,
                demand.product_name,
                demand.product_code,
                demand.new_description,
                demand.quantity,
                demand.required_by,
                demand.name,
                demand.contact_number,
            ),
        )
        conn.commit()
        cursor.execute("SELECT * FROM demands WHERE id = ?", (cursor.lastrowid,))
        return DemandResponse(**dict(cursor.fetchone()))


@app.get(
    "/demands", response_model=List[DemandResponse], dependencies=[Depends(verify_auth)]
)
def get_demands(
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    with get_db() as conn:
        cursor = conn.cursor()
        query = "SELECT * FROM demands WHERE 1=1"
        params = []

        if category:
            query += " AND category = ?"
            params.append(category)
        if date_from:
            query += " AND date(timestamp) >= ?"
            params.append(date_from)
        if date_to:
            query += " AND date(timestamp) <= ?"
            params.append(date_to)

        query += " ORDER BY timestamp DESC"
        cursor.execute(query, params)
        return [DemandResponse(**dict(row)) for row in cursor.fetchall()]


@app.put(
    "/demands/{demand_id}",
    response_model=DemandResponse,
    dependencies=[Depends(verify_auth)],
)
def update_demand(demand_id: int, update: DemandUpdate):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE demands SET status = ? WHERE id = ?", (update.status, demand_id)
        )
        conn.commit()
        cursor.execute("SELECT * FROM demands WHERE id = ?", (demand_id,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Demand not found")
        return DemandResponse(**dict(row))


@app.delete(
    "/demands/{demand_id}",
    dependencies=[Depends(verify_auth)],
)
def delete_demand(demand_id: int):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM demands WHERE id = ?", (demand_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="Demand not found")
        cursor.execute("DELETE FROM demands WHERE id = ?", (demand_id,))
        conn.commit()
        return {"success": True}


@app.post("/feedback", response_model=FeedbackResponse)
def create_feedback(feedback: FeedbackCreate):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO feedback (rating, text) VALUES (?, ?)",
            (feedback.rating, feedback.text),
        )
        conn.commit()
        cursor.execute("SELECT * FROM feedback WHERE id = ?", (cursor.lastrowid,))
        return FeedbackResponse(**dict(cursor.fetchone()))


@app.get(
    "/feedback",
    response_model=List[FeedbackResponse],
    dependencies=[Depends(verify_auth)],
)
def get_feedback(
    rating: Optional[int] = Query(None, ge=1, le=5),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    with get_db() as conn:
        cursor = conn.cursor()
        query = "SELECT * FROM feedback WHERE 1=1"
        params = []

        if rating:
            query += " AND rating = ?"
            params.append(rating)
        if date_from:
            query += " AND date(timestamp) >= ?"
            params.append(date_from)
        if date_to:
            query += " AND date(timestamp) <= ?"
            params.append(date_to)

        query += " ORDER BY timestamp DESC"
        cursor.execute(query, params)
        return [FeedbackResponse(**dict(row)) for row in cursor.fetchall()]


@app.get("/feedback/stats", dependencies=[Depends(verify_auth)])
def get_feedback_stats():
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT AVG(rating), COUNT(*) FROM feedback")
        row = cursor.fetchone()
        return {"average_rating": round(row[0], 2) if row[0] else 0, "count": row[1]}


@app.delete("/feedback/{feedback_id}", dependencies=[Depends(verify_auth)])
def delete_feedback(feedback_id: int):
    print(f"Attempting to delete feedback with ID: {feedback_id}")

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM feedback WHERE id = ?", (feedback_id,))
        row = cursor.fetchone()

        print(f"Found feedback row: {row}")

        if not row:
            print(f"Feedback with ID {feedback_id} not found")
            raise HTTPException(status_code=404, detail="Feedback not found")

        cursor.execute("DELETE FROM feedback WHERE id = ?", (feedback_id,))
        conn.commit()
        print(f"Successfully deleted feedback with ID: {feedback_id}")

    return {"success": True}


@app.get("/feedback/debug", dependencies=[Depends(verify_auth)])
def debug_feedback():
    """Debug endpoint to check existing feedback entries"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, rating, timestamp FROM feedback ORDER BY id")
        rows = cursor.fetchall()
        return {"feedback": [dict(row) for row in rows]}


@app.get("/")
def read_root():
    return "Yay! The backend is working!"


# Ensure images directory exists before mounting for static access
os.makedirs(IMAGES_DIR, exist_ok=True)

# Serve uploaded product images
app.mount("/images", StaticFiles(directory=IMAGES_DIR), name="images")


@app.post("/analyze-transactions")
async def analyze_transactions(
    file: UploadFile = File(...),
    date_from: Optional[str] = Form(None),
    date_to: Optional[str] = Form(None),
):
    try:
        if not file.filename.endswith(".csv"):
            raise HTTPException(status_code=400, detail="Only CSV files are allowed")

        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents))

        required_columns = [
            "Card Issuer",
            "Card Type",
            "Card Network",
            "Customer Payment Mode ID",
            "Name",
            "Transaction Date",
            "Transaction Amount",
            "Transaction Status",
        ]
        missing = [c for c in required_columns if c not in df.columns]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing)}",
            )

        # Preprocessing
        columns_to_drop = [
            "System",
            "Zone",
            "Store Name",
            "Store Address",
            "City",
            "POS",
            "Hardware Model",
            "Hardware ID",
            "Acquirer",
            "TID",
            "MID",
            "Is Emi",
            "Contactless",
            "Currency",
            "Transaction Id",
            "Sub Report Type",
            "Batch No",
            "Batch Status",
        ]
        df = df.drop(columns=columns_to_drop, errors="ignore")

        # Fill NA values
        df["Card Issuer"] = df["Card Issuer"].fillna("UPI")
        df["Card Type"] = df["Card Type"].fillna("UPI")
        df["Card Network"] = df["Card Network"].fillna("UPI")
        df["Customer Payment Mode ID"] = df["Customer Payment Mode ID"].fillna("NA")
        df["Name"] = df["Name"].fillna("NA")

        # convert datetime
        # Use dayfirst=True if your CSV dates are in DD-MM-YYYY or DD/MM/YYYY format
        df["Transaction Date"] = pd.to_datetime(df["Transaction Date"])

        # Store full range before filtering for the UI's filter component
        full_min_date = str(df["Transaction Date"].min().date())
        full_max_date = str(df["Transaction Date"].max().date())

        # Apply date filters if provided
        if date_from:
            df = df[df["Transaction Date"] >= pd.to_datetime(date_from)]
        if date_to:
            # Append end of day time to include all transactions on the selected 'to' day
            df = df[df["Transaction Date"] <= pd.to_datetime(date_to + " 23:59:59")]

        # Clean and convert Transaction Amount to numeric
        if "Transaction Amount" in df.columns:
            # Remove commas if any and convert to float
            df["Transaction Amount"] = pd.to_numeric(
                df["Transaction Amount"].astype(str).str.replace(",", ""),
                errors="coerce",
            ).fillna(0)

        # Analysis I & II: Separate High Value and Frequent Users
        grouped = df.groupby("Customer Payment Mode ID")

        high_value_users = []
        frequent_users = []

        for mode_id, group in grouped:
            high_val_txns = group[group["Transaction Amount"] >= 15000]

            user_info = {
                "name": str(group["Name"].iloc[0]),
                "payment_mode_id": str(mode_id),
                "card_type": str(group["Card Type"].iloc[0]),
                "total_transactions": len(group),
                "high_value_count": len(high_val_txns),
                "transactions": group[
                    ["Transaction Date", "Transaction Amount", "Transaction Status"]
                ]
                .astype(str)
                .to_dict("records"),
                "high_value_transactions": high_val_txns[
                    ["Transaction Date", "Transaction Amount", "Transaction Status"]
                ]
                .astype(str)
                .to_dict("records"),
            }

            if len(high_val_txns) > 0:
                high_value_users.append(user_info)

            if len(group) >= 2 and user_info["name"] != "NA":
                frequent_users.append(user_info)

        # Analysis III - Details of session expired / User Cancelled
        expired_cancelled = df[
            df["Transaction Status"].isin(["session expired", "User Cancelled"])
        ]
        expired_list = (
            expired_cancelled[
                [
                    "Customer Payment Mode ID",
                    "Name",
                    "Transaction Amount",
                    "Transaction Date",
                    "Transaction Status",
                ]
            ]
            .astype(str)
            .to_dict("records")
        )

        # Analysis IV - Suspected Rapid Use within 2 Mins
        df = df.sort_values(["Customer Payment Mode ID", "Transaction Date"])
        df["prev_time"] = df.groupby("Customer Payment Mode ID")[
            "Transaction Date"
        ].shift(1)
        df["time_diff_sec"] = (
            df["Transaction Date"] - df["prev_time"]
        ).dt.total_seconds()

        suspicious_rapid_use = df[
            (df["time_diff_sec"] < 120) & (df["Transaction Status"] == "Success")
        ]
        rapid_use_list = (
            suspicious_rapid_use[
                [
                    "Customer Payment Mode ID",
                    "Name",
                    "Transaction Amount",
                    "Transaction Date",
                    "time_diff_sec",
                ]
            ]
            .astype(str)
            .to_dict("records")
        )

        # Stats
        stats = {}
        for ct in ["DEBIT", "CREDIT", "UPI"]:
            filtered = df[df["Card Type"] == ct]
            stats[ct] = {
                "count": int(len(filtered)),
                "amount": float(filtered["Transaction Amount"].sum()),
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

    return {
        "date_range": {
            "start": str(df["Transaction Date"].min().date())
            if not df.empty
            else date_from,
            "end": str(df["Transaction Date"].max().date())
            if not df.empty
            else date_to,
        },
        "available_range": {"min": full_min_date, "max": full_max_date},
        "high_value_analysis": high_value_users,
        "frequent_analysis": frequent_users,
        "expired_cancelled": expired_list,
        "rapid_use": rapid_use_list,
        "stats": stats,
        "total_revenue": float(df["Transaction Amount"].sum()) if not df.empty else 0,
        "total_transactions": int(len(df)),
    }
